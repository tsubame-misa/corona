import openpyxl
import json

load_book = openpyxl.load_workbook('./data/08_地方公共団体別事業一覧（第2次実施計画分）九州沖縄.xlsx')
sheets = load_book['Sheet1']
path_w = './json_data/kyushu.json'

Datas = []

for i in range(2, 10530+1):
    prefecture = sheets.cell(row=i, column=1).value
    city = sheets.cell(row=i, column=2).value
    municipal_code = sheets.cell(row=i, column=3).value
    id = sheets.cell(row=i, column=4).value
    auxiliary_single = sheets.cell(row=i, column=5).value
    casebook_case_number = sheets.cell(row=i, column=6).value
    name = sheets.cell(row=i, column=7).value
    aim = sheets.cell(row=i, column=8).value
    detail = sheets.cell(row=i, column=9).value
    target = sheets.cell(row=i, column=10).value
    relationship = sheets.cell(row=i, column=11).value
    classification = sheets.cell(row=i, column=12).value
    start = sheets.cell(row=i, column=13).value
    end = sheets.cell(row=i, column=14).value
    cost = sheets.cell(row=i, column=15).value

    data = {
        "id": id,
        "県名": prefecture,
        "市町村名": city,
        "自治体コード": municipal_code,
        "補助・単独": auxiliary_single,
        "事例番号": casebook_case_number,
        "事業名": name,
        "目的": aim,
        "内容": detail,
        "対象": target,
        "関係": relationship,
        "区分": classification,
        "始期": start,
        "終期": end,
        "総事業費": cost,
    }

    Datas.append(data)

    if i % 100 == 0:
        print(i)

with open(path_w, mode='a', encoding='utf-8') as f:
    f.write(json.dumps(Datas, sort_keys=False, ensure_ascii=False, indent=4))

print("完了")
